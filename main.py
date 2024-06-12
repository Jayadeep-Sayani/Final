import copy
import csv
import random
import openpyxl

courses = []
people = []

class Course:

    prereqs = []
    simult_blocks = []
    nonsimulat_blocks = []
    sections = []
    
    def __init__(self, n, id, nt, em) -> None:
        self.course_name = n
        self.course_id = id
        self.num_terms = nt
        self.enrollment_max = em
        self.sections = [None] * self.num_terms
        for i in range(len(self.sections)):
            self.sections[i] = CourseSection(self, i)

    def add_student(self, new_student):
        for i in range(len(self.sections)):
            if len(self.sections[i].get_students()) < self.enrollment_max:
                self.sections[i].add_student(new_student)
                break
    
    def get_num_sections(self):
        return self.num_terms
    
    def get_section(self, i):
        return self.sections[i]
    
    def reset_sections(self):
        for i in range(len(self.sections)):
            self.sections[i] = CourseSection(self, i)

class CourseSection:
    def __init__(self, course, num) -> None:
        self.course = course
        self.block = -1
        self.index = -1
        self.section_num = num
        self.students = []
    
    def add_student(self, new_student):
        if len(self.students) < self.course.enrollment_max:
            if len(new_student.t.schedule[self.block]) > 0:
                return
            self.students.append(new_student)
            new_student.t.schedule[self.block].append(self)

    def set_block(self, b):
        self.block = b

    def set_index(self, i):
        self.index = i

    def get_course(self):
        return self.course
    
    def get_students(self):
        return self.students


class Student:
    def __init__(self, id) -> None:
        self.id = int(id)
        self.main_courses = []
        self.alternate_courses = []
        self.t = Timetable()
    
    def give_courses(self):
        for i in range(len(self.main_courses)):
            if self.main_courses[i] is not None:
                self.main_courses[i].add_student(self)
        for i in range(min(len(self.main_courses) - len(self.t.get_all_course_sections()), len(self.alternate_courses))):
            self.alternate_courses[i].add_student(self)
    
    def get_timetable(self):
        return self.t
    
    def get_main_courses(self):
        return self.main_courses


class Timetable:
    def __init__(self) -> None:
        self.schedule = [[], [], [], [], [], [], [], []]
    
    def get_all_course_sections(self):
        a = []
        for i in range(8):
            for cs in self.schedule[i]:
                a.append(cs)
        return a
    
    def add_section(self, slot, sec):
        self.schedule[slot].append(sec)

    def get_schedule(self, i):
        return self.schedule[i]
    
    def clone(self):
        new_ttble = Timetable()

        for i in range(8):
            for course_sec in self.schedule[i]:
                new_ttble.add_section(i, course_sec)

        return new_ttble


def generate_data():
    generate_course_data(filepath="data/Course Information.csv")
    generate_student_requests(filepath="data/Cleaned Student Requests.csv")
    generate_sequences()
    # generate_rules(pathone="data/Course Sequencing Rules", pathtwo="data/Course Blocking Rules")

def generate_course_data(filepath):
    course_data  = []

    with open(filepath, mode='r', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        for line in csv_reader:
            course_data.append(line)

        for course in course_data:
            course_name = course[1]
            course_id = course[0]
            course_num_terms = int(course[8])
            course_enrollment_max = int(course[7])

            course_class = Course(course_name, course_id, course_num_terms, course_enrollment_max)
            courses.append(course_class)


def generate_student_requests(filepath):
    student_data  = []

    with open(filepath, mode='r', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        for line in csv_reader:
            student_data.append(line)

        student_obj = None
        course = None
        for student in student_data:
            if student[0] == "ID":
                if student_obj is not None:
                    people.append(student_obj)
                student_obj = Student(student[1])
            else:
                course = get_course_data(student[0])
                if course is not None:
                    if student[3] == 'Y':
                        student_obj.alternate_courses.append(course)
                    elif student[3] == 'N':
                        student_obj.main_courses.append(course)

def generate_sequences():
    data = []
    with open("data/Course Sequencing Rules.csv", newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            data.append(row)

    for row in data:
        courseZero = None
        for cr in courses:
            if cr.course_id == row[0]:
                courseZero = cr
                
        # populate CourBefore when needed
        for code in row[1:]:
            for c in courses:
                if c.course_id == code and courseZero is not None:
                    c.prereqs.append(courseZero)

def generate_timetable():
    ttble = Timetable()

    for c in courses:
        for i in range(c.get_num_sections()):
            slot = random.randint(0, 7)
            ttble.add_section(slot, c.get_section(i))
            c.get_section(i).set_block(slot)
            c.get_section(i).set_index(len(ttble.get_schedule(slot)))

    return ttble

def get_course_data(code):
    return next((c for c in courses if c.course_id == code), None)


def create_timetables():
    ttble = generate_timetable()

    for student in people:
        student.give_courses()
    

    best_ttble = ttble.clone()
    high_score = requested_course_metrics()
    cur_score = 0

    for i in range(10):
        print(i)

        for course in courses:
            course.reset_sections()

        t = generate_timetable()

        generate_student_requests("data/Cleaned Student Requests.csv")

        for student in people:
            student.give_courses()
        
        cur_score = requested_course_metrics()

        if cur_score > high_score:
           best_ttble = t.clone()
           high_score = requested_course_metrics()
        

    create_master_table(best_ttble)
    print("Percent of all requested courses placed: ", f"{requested_course_metrics() * 100:.2f}", "%")
    print("Percent of people with 8/8 requested courses ", f"{requested_course_metrics_eight() * 100:.2f}", "%")
    print("Percent of people with 7-8/8 requested courses ", f"{requested_course_metrics_seven_eight() * 100:.2f}", "%")
    print("Percent of people with 8/8 requested or alternate courses ", f"{eight_req_or_alt() * 100:.2f}", "%")

def create_master_table(timetable):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Timetable"

    # Add headers
    ws.cell(row=1, column=1).value = "S1A"
    ws.cell(row=1, column=2).value = "S1B"
    ws.cell(row=1, column=3).value = "S1C"
    ws.cell(row=1, column=4).value = "S1D"
    ws.cell(row=1, column=5).value = "S2A"
    ws.cell(row=1, column=6).value = "S2B"
    ws.cell(row=1, column=7).value = "S2C"
    ws.cell(row=1, column=8).value = "S2D"

    # Populate data
    row = 2
    for slot in range(8):
        for course_section in timetable.get_schedule(slot):
            course = course_section.get_course()
            course_id = course.course_id
            students_enrolled = len(course_section.get_students())
            ws.cell(row=row, column=slot+1).value = f"{course_id} ({students_enrolled})"
            row += 1
        row = 2

    # Save the workbook
    wb.save("Master_Timetable.xlsx")



def requested_course_metrics():
        totalReqCourses = 0
        totalPlacedReqCourses = 0

        for student in people:

            totalReqCourses += len(student.get_main_courses())
            for req_course in student.get_main_courses():
                for act_course in student.get_timetable().get_all_course_sections():
                    if req_course.course_id == act_course.get_course().course_id:
                        totalPlacedReqCourses += 1
                        break

        return totalPlacedReqCourses / totalReqCourses

def requested_course_metrics_eight():
    num_students_with_all_requested_courses = 0

    for student in people:
        if len(student.get_main_courses()) == 8:  # Check if the student requested 8 main courses
            num_requested_courses_in_timetable = 0
            for req_course in student.get_main_courses():
                for act_course in student.get_timetable().get_all_course_sections():
                    if req_course.course_id == act_course.get_course().course_id:
                        num_requested_courses_in_timetable += 1
                        break
            if num_requested_courses_in_timetable == 8:  # Check if all 8 requested courses are in the timetable
                num_students_with_all_requested_courses += 1

    return num_students_with_all_requested_courses / len(people)

def requested_course_metrics_seven_eight():
    num_students_with_all_requested_courses = 0

    for student in people:
        if len(student.get_main_courses()) >= 7:  
            num_requested_courses_in_timetable = 0
            for req_course in student.get_main_courses():
                for act_course in student.get_timetable().get_all_course_sections():
                    if req_course.course_id == act_course.get_course().course_id:
                        num_requested_courses_in_timetable += 1
                        break
            if num_requested_courses_in_timetable >= 7:  # Check if all 8 requested courses are in the timetable
                num_students_with_all_requested_courses += 1

    return num_students_with_all_requested_courses / len(people)

def eight_req_or_alt():
    num_students_with_all_requested_courses = 0

    for student in people:
        if len(student.get_main_courses()) == 8:  # Check if the student requested 8 courses in total
            num_requested_courses_in_timetable = 0
            for req_course in student.get_main_courses():
                for act_course in student.get_timetable().get_all_course_sections():
                    if req_course.course_id == act_course.get_course().course_id:
                        num_requested_courses_in_timetable += 1
                        break
            for alt_course in student.alternate_courses:
                for act_course in student.get_timetable().get_all_course_sections():
                    if alt_course.course_id == act_course.get_course().course_id:
                        num_requested_courses_in_timetable += 1
                        break
            if num_requested_courses_in_timetable == 8:  # Check if all 8 requested or alternate courses are in the timetable
                num_students_with_all_requested_courses += 1

    return num_students_with_all_requested_courses / len(people)


if __name__ == '__main__':
    generate_data()
    create_timetables()